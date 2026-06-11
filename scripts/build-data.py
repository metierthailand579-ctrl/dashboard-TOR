#!/usr/bin/env python3
"""แปลง โครงสร้างโฟลเดอร์_รวม_TOR.xlsx -> data/projects.json + data/overview.json

รัน: python3 scripts/build-data.py
อ้างอิง skill: spec-miner (สกัดข้อมูลจากเอกสารต้นทาง)
"""
import json
import os
import re
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "โครงสร้างโฟลเดอร์_รวม_TOR.xlsx")
OUT = os.path.join(ROOT, "data")


def slug_files(cell):
    return [f.strip() for f in str(cell or "").split("\n") if f.strip()]


def build_projects(ws, files_by_code):
    """ชีตหลัก (8 คอลัมน์):
    ลำดับ|ประเภท|ช่วงวงเงิน|รหัส|ชื่อ|จำนวนไฟล์|กลุ่มหลัก|กลุ่มย่อย
    รายชื่อไฟล์ดึงจากชีต "เช็คการอ่านไฟล์" ผ่าน files_by_code
    กลุ่มเป็นชุดเดียว: 7 หมวดงาน (ขึ้นต้นด้วยเลข) + โอกาส Metier (Software/Creative/Media)
    Metier = derive จากกลุ่มหลักที่ไม่ขึ้นต้นด้วยตัวเลข
    """
    rows = list(ws.iter_rows(values_only=True))
    projects = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        code = str(r[3]).strip()
        files = files_by_code.get(code, [])
        ftypes = sorted({os.path.splitext(f)[1].lower() for f in files if "." in f})
        tor_files = [f for f in files if re.search(r"tor", f, re.IGNORECASE)]

        work_group = (r[6] or "").strip() if r[6] else "ไม่ระบุ"
        work_sub = (r[7] or "").strip() if r[7] else "ไม่ระบุ"
        # โอกาส Metier = กลุ่มหลักที่ไม่ขึ้นต้นด้วยเลข (Software/Creative/Media Management)
        is_metier = bool(work_group) and not re.match(r"^\d", work_group)
        metier_group = work_group if is_metier else "NOT_APPLICABLE"
        metier_sub = work_sub if is_metier else "NOT_APPLICABLE"

        projects.append({
            "order": r[0],
            "type": (r[1] or "").strip(),
            "budgetRange": (r[2] or "").strip(),
            "code": code,
            "name": (r[4] or "").strip(),
            # group/subGroup = กลุ่มงาน (ระบบ A) — คงชื่อเดิมเพื่อ backward-compat
            "group": work_group,
            "subGroup": work_sub,
            "workGroup": work_group,
            "workSubGroup": work_sub,
            "metierGroup": metier_group,
            "metierSubGroup": metier_sub,
            "fileCount": int(r[5]) if isinstance(r[5], (int, float)) else len(files),
            "fileTypes": ftypes,
            "files": files,
            "torFiles": tor_files,
        })
    return projects


# ---- งานจัดทำ TOR: อ่านจากชีต "TOR" --------------------------------------
# สถานะการจัดทำ TOR (workflow) — เลือก/บันทึกฝั่ง browser (localStorage)
TOR_STATUSES = ["to do", "skill.md", "วางโครงร่างTOR", "TOR เสร็จสมบูรณ์"]
TOR_DEFAULT_STATUS = "to do"


def build_tor(ws):
    """ชีต "TOR" (3 คอลัมน์): โครงการ | จำนวน TOR | Status
    คืนรายการงานจัดทำ TOR พร้อมสถานะเริ่มต้น
    (Status ใน Excel ว่าง → ใช้ค่าเริ่มต้น "to do"; ผู้ใช้เปลี่ยน/บันทึกในเว็บ)
    """
    items = []
    order = 0
    for r in list(ws.iter_rows(values_only=True))[1:]:
        name = str(r[0]).strip() if r[0] is not None else ""
        if not name:
            continue
        order += 1
        raw_status = str(r[2]).strip() if r[2] is not None else ""
        status = raw_status if raw_status in TOR_STATUSES else TOR_DEFAULT_STATUS
        items.append({
            "order": order,
            "name": name,
            "torCount": int(r[1]) if isinstance(r[1], (int, float)) else 0,
            "status": status,
        })
    return items


def build_overview(ws):
    """ชีตภาพรวม: ประเภท | ช่วงวงเงิน | จำนวนโครงการ | จำนวนไฟล์"""
    rows = list(ws.iter_rows(values_only=True))
    items = []
    current_type = None
    for r in rows:
        a, b, c, d = (r + (None,) * 4)[:4]
        if not isinstance(c, (int, float)):
            # หัวข้อ/หัวตาราง/บรรทัดว่าง
            if a and isinstance(a, str) and a.startswith("ประเภท"):
                continue
            continue
        label = (a or "").strip() if isinstance(a, str) else ""
        if label.startswith("รวม"):
            continue  # ข้ามแถวรวม คำนวณเองทีหลัง
        if label and not label.startswith("รวม"):
            current_type = label
        items.append({
            "type": current_type,
            "budgetRange": (b or "").strip() if isinstance(b, str) else "",
            "projectCount": int(c),
            "fileCount": int(d) if isinstance(d, (int, float)) else 0,
        })
    return items


# ---- TOR Health: อ่านผลตรวจการอ่านไฟล์จากชีต "เช็คการอ่านไฟล์" -------------
# สถานะรายไฟล์ (ตามที่บันทึกใน Excel): อ่านได้ / ต้อง OCR / อ่านไม่ได้
ST_READ = "อ่านได้"
ST_OCR = "ต้อง OCR"
ST_UNREAD = "อ่านไม่ได้"
ST_PARTIAL = "อ่านได้บางส่วน"
ST_NONE = "ไม่มีข้อมูล"

# ผล OCR (คอลัมน์ใหม่)
OCR_OK = "OCR อ่านได้"
OCR_FAIL = "OCR อ่านไม่ได้"

# สถานะการเข้าถึงข้อความ (รวมผล OCR แล้ว)
AC_ALL = "เข้าถึงได้ทั้งหมด"
AC_PARTIAL = "เข้าถึงได้บางส่วน"
AC_NONE = "เข้าถึงไม่ได้"


def file_accessible(status, ocr):
    """ไฟล์เข้าถึงข้อความได้หรือไม่ (มี text อยู่แล้ว หรือ OCR สำเร็จ)"""
    return status == ST_READ or ocr == OCR_OK


def access_status(accessible, total):
    if total == 0:
        return ST_NONE
    if accessible == total:
        return AC_ALL
    if accessible > 0:
        return AC_PARTIAL
    return AC_NONE


def rollup_status(readable, ocr, unreadable):
    """สรุปสถานะระดับโครงการจากจำนวนไฟล์แต่ละสถานะ"""
    total = readable + ocr + unreadable
    if total == 0:
        return ST_NONE
    if readable == total:
        return ST_READ
    if readable > 0:
        return ST_PARTIAL
    if ocr > 0:
        return ST_OCR
    return ST_UNREAD


def summary_text(readable, ocr, unreadable):
    parts = []
    if readable:
        parts.append(f"{ST_READ}:{readable}")
    if ocr:
        parts.append(f"{ST_OCR}:{ocr}")
    if unreadable:
        parts.append(f"{ST_UNREAD}:{unreadable}")
    return " / ".join(parts) or ST_NONE


def build_health(ws):
    """ชีตเช็คการอ่านไฟล์ (10 คอลัมน์):
    ลำดับ|ประเภท|ช่วงวงเงิน|โครงการ(code)|ชื่อไฟล์|ชนิด|ขนาด(KB)|สถานะการอ่าน|ผล OCR|รายละเอียด
    คืน (health dict {code: ...}, file_summary dict)
    """
    by_code = {}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        code = str(r[3]).strip() if r[3] is not None else ""
        if not code:
            continue
        status = (r[7] or "").strip()
        ocr = (r[8] or "").strip() if r[8] else ""
        by_code.setdefault(code, []).append({
            "name": (r[4] or "").strip(),
            "status": status,
            "ocr": ocr,
            "sizeKB": round(float(r[6]), 1) if isinstance(r[6], (int, float)) else None,
            "detail": (r[9] or "").strip() if r[9] else "",
            "accessible": file_accessible(status, ocr),
        })

    health = {}
    file_summary = {"textReady": 0, "ocrOk": 0, "ocrFail": 0, "corrupt": 0,
                    "accessible": 0, "total": 0}
    for code, files in by_code.items():
        readable = sum(1 for f in files if f["status"] == ST_READ)
        ocr_need = sum(1 for f in files if f["status"] == ST_OCR)
        unreadable = sum(1 for f in files if f["status"] == ST_UNREAD)
        ocr_ok = sum(1 for f in files if f["ocr"] == OCR_OK)
        ocr_fail = sum(1 for f in files if f["ocr"] == OCR_FAIL)
        accessible = sum(1 for f in files if f["accessible"])
        total = len(files)

        # สรุปการเข้าถึงข้อความระดับไฟล์ (ทั้งระบบ)
        file_summary["textReady"] += readable
        file_summary["ocrOk"] += ocr_ok
        file_summary["ocrFail"] += ocr_fail
        file_summary["corrupt"] += unreadable
        file_summary["accessible"] += accessible
        file_summary["total"] += total

        health[code] = {
            "status": rollup_status(readable, ocr_need, unreadable),
            "summary": summary_text(readable, ocr_need, unreadable),
            "accessStatus": access_status(accessible, total),
            "accessSummary": f"เข้าถึงข้อความได้ {accessible}/{total} ไฟล์",
            "counts": {
                "readable": readable,
                "ocr": ocr_need,
                "unreadable": unreadable,
                "ocrOk": ocr_ok,
                "ocrFail": ocr_fail,
                "accessible": accessible,
                "total": total,
            },
            "files": files,
        }
    return health, file_summary


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    if "เช็คการอ่านไฟล์" in wb.sheetnames:
        health, file_summary = build_health(wb["เช็คการอ่านไฟล์"])
    else:
        health, file_summary = {}, {"textReady": 0, "ocrOk": 0, "ocrFail": 0,
                                    "corrupt": 0, "accessible": 0, "total": 0}
    files_by_code = {code: [f["name"] for f in h["files"]] for code, h in health.items()}
    projects = build_projects(wb["รายการโครงการทั้งหมด"], files_by_code)
    overview = build_overview(wb["ภาพรวม"])
    tor = build_tor(wb["TOR"]) if "TOR" in wb.sheetnames else []

    # โครงการที่ไม่มีข้อมูลตรวจ → สถานะ "ไม่มีข้อมูล"
    none_health = {"status": ST_NONE, "summary": ST_NONE,
                   "accessStatus": ST_NONE, "accessSummary": ST_NONE,
                   "counts": {"readable": 0, "ocr": 0, "unreadable": 0,
                              "ocrOk": 0, "ocrFail": 0, "accessible": 0, "total": 0},
                   "files": []}
    for p in projects:
        health.setdefault(p["code"], dict(none_health))

    # สรุประดับโครงการ (จำนวนโครงการต่อสถานะ) — ใช้ทำชิปฟิลเตอร์
    project_summary = {}
    for code in (p["code"] for p in projects):
        st = health[code]["status"]
        project_summary[st] = project_summary.get(st, 0) + 1

    os.makedirs(OUT, exist_ok=True)
    with open(os.path.join(OUT, "projects.json"), "w", encoding="utf-8") as f:
        json.dump(projects, f, ensure_ascii=False, indent=2)
    with open(os.path.join(OUT, "overview.json"), "w", encoding="utf-8") as f:
        json.dump(overview, f, ensure_ascii=False, indent=2)
    with open(os.path.join(OUT, "tor.json"), "w", encoding="utf-8") as f:
        json.dump(tor, f, ensure_ascii=False, indent=2)
    with open(os.path.join(OUT, "healthcheck.json"), "w", encoding="utf-8") as f:
        json.dump({
            "engine": "excel:เช็คการอ่านไฟล์",
            "summary": project_summary,
            "fileSummary": file_summary,
            "projects": {p["code"]: health[p["code"]] for p in projects},
        }, f, ensure_ascii=False, indent=2)

    types = sorted(set(p["type"] for p in projects))
    budgets = sorted(set(p["budgetRange"] for p in projects))
    print(f"✓ projects.json: {len(projects)} โครงการ")
    print(f"✓ overview.json: {len(overview)} แถว")
    print(f"✓ tor.json: {len(tor)} งานจัดทำ TOR")
    print(f"✓ healthcheck.json: สรุปต่อโครงการ {project_summary}")
    print(f"  เข้าถึงข้อความได้ {file_summary['accessible']}/{file_summary['total']} ไฟล์ "
          f"(text {file_summary['textReady']} + OCR {file_summary['ocrOk']})")
    print(f"  ประเภท: {types}")
    print(f"  ช่วงวงเงิน: {len(budgets)} ช่วง")


if __name__ == "__main__":
    main()
