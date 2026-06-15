#!/usr/bin/env python3
"""apply_canonical_subgroups.py — แก้กลุ่มย่อย Metier ใน Excel เป็น canonical sub-service

taxonomy ทางการ (จาก dropdown UI ของ Metier) — กลุ่มหลัก (col7) คงเดิม "X Metier"
แก้เฉพาะกลุ่มย่อย (col8) ของโครงการ Metier ตาม MAP (จับเข้าช่องใกล้สุดรายตัว)
ที่ไม่มีช่องตรง → "อื่น ๆ (ยังไม่จัดประเภท)"

รัน: python3 scripts/apply_canonical_subgroups.py
"""
import os
import shutil
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "โครงสร้างโฟลเดอร์_รวม_TOR.xlsx")
BACKUP = os.path.join(ROOT, "_tmp_before_canonical.xlsx")

OTHER = "อื่น ๆ (ยังไม่จัดประเภท)"

# order -> canonical sub-service (กลุ่มหลักไม่เปลี่ยน)
MAP = {
    # ── Software Metier ──
    74: "ERP/CRM Systems", 121: "ERP/CRM Systems", 124: "ERP/CRM Systems",
    134: "ERP/CRM Systems", 151: "ERP/CRM Systems", 157: "ERP/CRM Systems",
    162: "ERP/CRM Systems", 177: "ERP/CRM Systems", 191: "ERP/CRM Systems",
    203: "ERP/CRM Systems", 206: "ERP/CRM Systems", 214: "ERP/CRM Systems",
    216: "ERP/CRM Systems", 221: "ERP/CRM Systems", 240: "ERP/CRM Systems",
    246: "ERP/CRM Systems", 287: "ERP/CRM Systems", 288: "ERP/CRM Systems",
    294: "ERP/CRM Systems", 306: "ERP/CRM Systems",
    160: "Software Quality Assurance & Cyber Security",
    253: "Software Quality Assurance & Cyber Security",
    226: "Mobile Application", 235: "Mobile Application",
    63: OTHER, 174: OTHER, 219: OTHER, 245: OTHER,
    # ── Creative Metier ──
    167: "Event Marketing", 173: "Event Marketing", 236: "Event Marketing",
    266: "Event Marketing", 276: "Event Marketing", 277: "Event Marketing",
    278: "Event Marketing", 296: "Event Marketing", 297: "Event Marketing",
    301: "Event Marketing",
    256: "Content Creation", 269: "Content Creation", 275: "Content Creation",
    123: "Graphic Design",
    131: OTHER,
    # ── Media Metier ──
    274: "Public Relations", 279: "Public Relations", 280: "Public Relations",
    281: "Public Relations",
    32: "Offline Media",
    # ── Marketing Metier ──
    127: "Brand Strategy",
}

GROUP_ORDER = [
    "1. บริหารจัดการ", "2. โครงสร้างพื้นฐาน", "3. คุณภาพชีวิต", "4. ชุมชนเข้มแข็ง",
    "5. การศึกษา ศาสนา กีฬา", "6. ความสะอาดและสิ่งแวดล้อม", "7. ครุภัณฑ์",
    "Software Metier", "Creative Metier", "Media Metier", "Marketing Metier",
]
METIER_GROUPS = {"Software Metier", "Creative Metier", "Media Metier", "Marketing Metier"}


def apply_mapping(ws):
    changed = []
    for row in ws.iter_rows(min_row=2):
        order = row[0].value
        if order in MAP:
            old = row[7].value
            row[7].value = MAP[order]      # col8 กลุ่มย่อย
            changed.append((order, old, MAP[order]))
    return changed


def rebuild_summary(wb, ws_p):
    breakdown = defaultdict(Counter)
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        g = (row[6] or "").strip()
        s = (row[7] or "").strip() or "ไม่ระบุ"
        if g:
            breakdown[g][s] += 1
    idx = wb.sheetnames.index("สรุปกลุ่ม")
    wb.remove(wb["สรุปกลุ่ม"])
    ws = wb.create_sheet("สรุปกลุ่ม", idx)
    bold = Font(bold=True)
    yellow = PatternFill("solid", fgColor="FFF6CC")
    ws.append(["สรุปการจัดกลุ่ม (ชุดเดียว) : กลุ่มหลัก → กลุ่มย่อย"]); ws["A1"].font = bold
    ws.append(["กลุ่มย่อย Metier ใช้ canonical sub-service (Software/Creative/Media/Marketing)"])
    ws.append([])
    ws.append(["กลุ่มหลัก", "กลุ่มย่อย", "จำนวนโครงการ"])
    for c in ws[ws.max_row]:
        c.font = bold
    grand = 0
    for g in GROUP_ORDER:
        subs = breakdown.get(g)
        is_metier = g in METIER_GROUPS
        if not subs:
            if is_metier:
                ws.append([g, "(ยังไม่มีโครงการในกลุ่ม)", 0])
                for c in ws[ws.max_row]:
                    c.fill = yellow
                ws.append([f"รวม {g}", "", 0])
                for c in ws[ws.max_row]:
                    c.font = bold
            continue
        first = True
        for s, n in sorted(subs.items(), key=lambda kv: (-kv[1], kv[0])):
            ws.append([g if first else "", s, n])
            if is_metier:
                for c in ws[ws.max_row]:
                    c.fill = yellow
            first = False
        total = sum(subs.values()); grand += total
        ws.append([f"รวม {g}", "", total])
        for c in ws[ws.max_row]:
            c.font = bold
    ws.append(["รวมทั้งหมด", "", grand])
    for c in ws[ws.max_row]:
        c.font = bold
    for col, w in (("A", 26), ("B", 42), ("C", 14)):
        ws.column_dimensions[col].width = w
    return breakdown


def main():
    shutil.copy(XLSX, BACKUP)
    wb = openpyxl.load_workbook(XLSX)
    ws_p = wb["รายการโครงการทั้งหมด"]
    changed = apply_mapping(ws_p)
    print(f"แก้กลุ่มย่อย {len(changed)} โครงการ")
    bd = rebuild_summary(wb, ws_p)
    for g in ["Software Metier", "Creative Metier", "Media Metier", "Marketing Metier"]:
        line = ", ".join(f"{s}:{n}" for s, n in sorted(bd[g].items(), key=lambda kv: -kv[1]))
        print(f"  {g}: {line}")
    wb.save(XLSX)
    os.remove(BACKUP)
    print(f"✓ บันทึก {os.path.basename(XLSX)} (git history เก็บเวอร์ชันก่อนแก้ไว้แล้ว)")


if __name__ == "__main__":
    main()
