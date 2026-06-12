#!/usr/bin/env python3
"""conform_to_skill.py — แก้กลุ่มย่อย Metier ใน Excel ให้ตรงกับ SKILL §3 (Layer 2)

ก่อนหน้านี้ Excel ใช้ชื่อกลุ่มย่อย Metier บางตัวที่ไม่อยู่ในรายการที่ SKILL นิยาม
(System Development, Cybersecurity/Certificate, Software/License, Content Creation /
Graphic Design รวมเป็นช่องเดียว, Marketing/Brand-Corporate-Image)
สคริปต์นี้ remap ให้ตรงชื่อ SKILL เดิม + Marketing → 0 (ตาม SKILL "อย่าฝืน promote")

แก้คอลัมน์ใน "รายการโครงการทั้งหมด": col7 กลุ่มหลัก, col8 กลุ่มย่อย (match ด้วย col1 ลำดับ)
แล้ว regenerate ชีต "สรุปกลุ่ม" ใหม่จากข้อมูลที่แก้แล้ว (ให้ Excel สอดคล้องกัน)

รัน: python3 scripts/conform_to_skill.py   (backup เป็น *_BEFORE_SKILL.xlsx)
"""
import os
import shutil
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "โครงสร้างโฟลเดอร์_รวม_TOR.xlsx")
BACKUP = os.path.join(ROOT, "โครงสร้างโฟลเดอร์_รวม_TOR_BEFORE_SKILL.xlsx")

# ── mapping: ลำดับโครงการ → (กลุ่มหลักใหม่, กลุ่มย่อยใหม่) ตาม SKILL §3 ───────────
SD = ("Software Metier", "Smart City/Platform Development")
MAP = {
    # System Development (19) → Smart City/Platform Development (default)
    74: SD, 121: SD, 124: SD, 134: SD, 151: SD, 162: SD, 177: SD, 191: SD,
    203: SD, 206: SD, 214: SD, 216: SD, 221: SD, 240: SD, 246: SD, 294: SD, 306: SD,
    226: ("Software Metier", "Mobile Application"),        # ...ผ่านมือถือ Mobile Check-Out
    219: ("Software Metier", "Server/IT Infrastructure"),  # ซ่อมแซมคอมพิวเตอร์
    # Cybersecurity/Certificate (3) → Server/IT Infrastructure
    160: ("Software Metier", "Server/IT Infrastructure"),
    245: ("Software Metier", "Server/IT Infrastructure"),  # FLAG: เดิมเป็นกิจกรรม PR ไซเบอร์
    253: ("Software Metier", "Server/IT Infrastructure"),
    # Software/License (1) → Server/IT Infrastructure
    63: ("Software Metier", "Server/IT Infrastructure"),   # FLAG: ลิขสิทธิ์ Canva
    # Content Creation / Graphic Design (5) → แยก
    123: ("Creative Metier", "Graphic Design"),            # สื่อสิ่งพิมพ์
    256: ("Creative Metier", "Content Creation"),
    269: ("Creative Metier", "Content Creation"),
    275: ("Creative Metier", "Content Creation"),
    278: ("Creative Metier", "Content Creation"),
    # Marketing → 0 (SKILL): ย้ายไป Media/Public Relations
    127: ("Media Metier", "Public Relations"),
}

# ลำดับการแสดงกลุ่มหลักในชีตสรุปกลุ่ม
GROUP_ORDER = [
    "1. บริหารจัดการ", "2. โครงสร้างพื้นฐาน", "3. คุณภาพชีวิต", "4. ชุมชนเข้มแข็ง",
    "5. การศึกษา ศาสนา กีฬา", "6. ความสะอาดและสิ่งแวดล้อม", "7. ครุภัณฑ์",
    "Software Metier", "Creative Metier", "Media Metier", "Marketing Metier",
]
METIER_GROUPS = {"Software Metier", "Creative Metier", "Media Metier", "Marketing Metier"}


def apply_mapping(ws):
    """แก้ col7/col8 ใน 'รายการโครงการทั้งหมด' ตาม MAP (header อยู่แถว 1)"""
    changed = []
    for row in ws.iter_rows(min_row=2):
        order = row[0].value
        if order in MAP:
            new_g, new_s = MAP[order]
            old_g, old_s = row[6].value, row[7].value
            row[6].value = new_g  # col7 กลุ่มหลัก
            row[7].value = new_s  # col8 กลุ่มย่อย
            changed.append((order, old_g, old_s, new_g, new_s))
    return changed


def rebuild_summary(wb, ws_projects):
    """สร้างชีต 'สรุปกลุ่ม' ใหม่จากข้อมูลที่แก้แล้ว (กลุ่มหลัก→กลุ่มย่อย→จำนวน)"""
    breakdown = defaultdict(Counter)
    for row in ws_projects.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        g = (row[6] or "").strip()
        s = (row[7] or "").strip() or "ไม่ระบุ"
        if g:
            breakdown[g][s] += 1

    idx = wb.sheetnames.index("สรุปกลุ่ม")
    wb.remove(wb["สรุปกลุ่ม"])
    ws = wb.create_sheet("สรุปกลุ่ม", idx)

    bold = Font(bold=True)
    yellow = PatternFill("solid", fgColor="FFF6CC")
    ws.append(["สรุปการจัดกลุ่ม (ชุดเดียว) : กลุ่มหลัก → กลุ่มย่อย"])
    ws["A1"].font = bold
    ws.append(["กลุ่มหลักรวม 7 หมวดงาน + Software/Creative/Media/Marketing Metier "
               "(กลุ่มย่อย Metier ตรงตาม SKILL thai-municipal-project-grouping §3)"])
    ws.append([])
    ws.append(["กลุ่มหลัก", "กลุ่มย่อย", "จำนวนโครงการ"])
    for c in ws[ws.max_row]:
        c.font = bold

    grand = 0
    for g in GROUP_ORDER:
        subs = breakdown.get(g)
        if not subs:
            # ระบุกลุ่ม Metier ที่ยังไม่มีโครงการไว้เสมอ (ตามคำขอ: Marketing ต้องอยู่ในระบบ)
            if g in METIER_GROUPS:
                ws.append([g, "(ยังไม่มีโครงการในกลุ่ม)", 0])
                for c in ws[ws.max_row]:
                    c.fill = yellow
                ws.append([f"รวม {g}", "", 0])
                for c in ws[ws.max_row]:
                    c.font = bold
            continue
        is_metier = g in METIER_GROUPS
        first = True
        for s, n in sorted(subs.items(), key=lambda kv: (-kv[1], kv[0])):
            ws.append([g if first else "", s, n])
            r = ws[ws.max_row]
            if is_metier:
                for c in r:
                    c.fill = yellow
            first = False
        total = sum(subs.values())
        grand += total
        ws.append([f"รวม {g}", "", total])
        for c in ws[ws.max_row]:
            c.font = bold
    ws.append(["รวมทั้งหมด", "", grand])
    for c in ws[ws.max_row]:
        c.font = bold

    metier_total = sum(sum(breakdown[g].values()) for g in METIER_GROUPS if g in breakdown)
    ws.append([])
    ws.append([f"➜ โครงการที่เป็นโอกาส Metier = {metier_total} จาก {grand} โครงการ"])

    for col, w in (("A", 26), ("B", 38), ("C", 14)):
        ws.column_dimensions[col].width = w
    return breakdown


def main():
    shutil.copy(XLSX, BACKUP)
    wb = openpyxl.load_workbook(XLSX)  # ไม่ data_only (ไม่มีสูตร — ค่าคงที่ทั้งหมด)
    ws_p = wb["รายการโครงการทั้งหมด"]

    changed = apply_mapping(ws_p)
    print(f"แก้ {len(changed)} โครงการใน 'รายการโครงการทั้งหมด':")
    for order, og, os_, ng, ns in changed:
        tag = "  ⚠️" if (og, os_) != (ng, ns) and ng != og else "   "
        print(f"{tag}[{order}] {os_}  →  {ns}" + (f"   ({og} → {ng})" if ng != og else ""))

    bd = rebuild_summary(wb, ws_p)
    print("\nสร้างชีต 'สรุปกลุ่ม' ใหม่ — กลุ่มย่อย Metier หลังแก้:")
    for g in ["Software Metier", "Creative Metier", "Media Metier", "Marketing Metier"]:
        subs = bd.get(g, {})
        line = ", ".join(f"{s}:{n}" for s, n in sorted(subs.items(), key=lambda kv: -kv[1]))
        print(f"  {g}: {line or '(0 — ตาม SKILL)'}")

    wb.save(XLSX)
    print(f"\n✓ บันทึก {os.path.basename(XLSX)} (backup: {os.path.basename(BACKUP)})")


if __name__ == "__main__":
    main()
